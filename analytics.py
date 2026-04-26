"""Wait-time session logging and aggregation, backed by SQLite (WAL mode).

Each camera service writes a row when a person's absence window expires
(they're done waiting). The login gateway reads aggregates back out for
the reports page and Excel exports.
"""

from __future__ import annotations

import os
import sqlite3
import time
from datetime import datetime
from io import BytesIO
from typing import List, Optional, Sequence


DEFAULT_DB_PATH = os.environ.get("DB_PATH", "/var/lib/ai-track/waittime.db")
# Sessions shorter than this are ignored (just passers-by, not real waits)
MIN_LOGGED_WAIT_S = float(os.environ.get("MIN_LOGGED_WAIT_S", "5.0"))


def _ensure_dir(db_path: str) -> None:
    parent = os.path.dirname(db_path)
    if parent:
        os.makedirs(parent, exist_ok=True)


def _connect(db_path: str = DEFAULT_DB_PATH) -> sqlite3.Connection:
    _ensure_dir(db_path)
    conn = sqlite3.connect(db_path, timeout=10.0, isolation_level=None)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.row_factory = sqlite3.Row
    return conn


def init_db(db_path: str = DEFAULT_DB_PATH) -> None:
    conn = _connect(db_path)
    try:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS wait_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                camera TEXT NOT NULL,
                label_id INTEGER NOT NULL,
                first_seen REAL NOT NULL,
                last_seen REAL NOT NULL,
                total_wait_s REAL NOT NULL,
                date TEXT NOT NULL,
                created_at REAL NOT NULL DEFAULT (unixepoch())
            );
            CREATE INDEX IF NOT EXISTS idx_camera_date ON wait_sessions(camera, date);
            CREATE INDEX IF NOT EXISTS idx_camera_first_seen ON wait_sessions(camera, first_seen);
            """
        )
    finally:
        conn.close()


def log_session(
    camera: str,
    label_id: int,
    first_seen: float,
    last_seen: float,
    total_wait_s: float,
    db_path: str = DEFAULT_DB_PATH,
) -> Optional[int]:
    """Insert a completed wait session. Returns the new row id, or None if filtered."""
    if total_wait_s < MIN_LOGGED_WAIT_S:
        return None
    date_str = datetime.fromtimestamp(first_seen).strftime("%Y-%m-%d")
    conn = _connect(db_path)
    try:
        cur = conn.execute(
            "INSERT INTO wait_sessions (camera, label_id, first_seen, last_seen, total_wait_s, date) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (camera, int(label_id), float(first_seen), float(last_seen), float(total_wait_s), date_str),
        )
        return cur.lastrowid
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Read queries
# ---------------------------------------------------------------------------


def list_sessions(camera: str, date: str, db_path: str = DEFAULT_DB_PATH) -> List[sqlite3.Row]:
    conn = _connect(db_path)
    try:
        return conn.execute(
            "SELECT id, label_id, first_seen, last_seen, total_wait_s "
            "FROM wait_sessions WHERE camera=? AND date=? ORDER BY first_seen",
            (camera, date),
        ).fetchall()
    finally:
        conn.close()


def daily_summary(
    camera: str,
    from_date: str,
    to_date: str,
    db_path: str = DEFAULT_DB_PATH,
) -> List[sqlite3.Row]:
    conn = _connect(db_path)
    try:
        return conn.execute(
            "SELECT date, COUNT(*) AS people, AVG(total_wait_s) AS avg_wait, "
            "MIN(total_wait_s) AS min_wait, MAX(total_wait_s) AS max_wait "
            "FROM wait_sessions WHERE camera=? AND date BETWEEN ? AND ? "
            "GROUP BY date ORDER BY date",
            (camera, from_date, to_date),
        ).fetchall()
    finally:
        conn.close()


def monthly_summary(camera: str, year: int, db_path: str = DEFAULT_DB_PATH) -> List[sqlite3.Row]:
    conn = _connect(db_path)
    try:
        return conn.execute(
            "SELECT substr(date,1,7) AS month, COUNT(*) AS people, AVG(total_wait_s) AS avg_wait "
            "FROM wait_sessions WHERE camera=? AND date BETWEEN ? AND ? "
            "GROUP BY month ORDER BY month",
            (camera, f"{year}-01-01", f"{year}-12-31"),
        ).fetchall()
    finally:
        conn.close()


def list_cameras(db_path: str = DEFAULT_DB_PATH) -> List[str]:
    conn = _connect(db_path)
    try:
        return [r[0] for r in conn.execute(
            "SELECT DISTINCT camera FROM wait_sessions ORDER BY camera"
        ).fetchall()]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Excel exports (openpyxl, lazy-imported so this module is usable without it)
# ---------------------------------------------------------------------------


def _fmt_hms(seconds: float) -> str:
    seconds = max(0, int(seconds))
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def make_daily_xlsx(camera: str, date: str, db_path: str = DEFAULT_DB_PATH) -> bytes:
    """One row per person waiting on that day. Columns: ID, first/last seen, wait."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = list_sessions(camera, date, db_path=db_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center")

    ws["A1"] = "Camera"
    ws["B1"] = camera
    ws["A2"] = "Date"
    ws["B2"] = date
    ws["A1"].font = bold
    ws["A2"].font = bold

    headers = ["#", "Person ID", "First seen", "Last seen", "Wait (seconds)", "Wait (mm:ss)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = center

    for i, r in enumerate(rows, 1):
        ws.cell(row=4 + i, column=1, value=i)
        ws.cell(row=4 + i, column=2, value=r["label_id"])
        ws.cell(row=4 + i, column=3, value=datetime.fromtimestamp(r["first_seen"]).strftime("%H:%M:%S"))
        ws.cell(row=4 + i, column=4, value=datetime.fromtimestamp(r["last_seen"]).strftime("%H:%M:%S"))
        ws.cell(row=4 + i, column=5, value=round(float(r["total_wait_s"]), 1))
        ws.cell(row=4 + i, column=6, value=_fmt_hms(float(r["total_wait_s"])))

    # Summary row
    summary_row = 4 + len(rows) + 2
    if rows:
        avg = sum(float(r["total_wait_s"]) for r in rows) / len(rows)
        ws.cell(row=summary_row, column=4, value="Total / Average:").font = bold
        ws.cell(row=summary_row, column=2, value=f"{len(rows)} people").font = bold
        ws.cell(row=summary_row, column=5, value=round(avg, 1)).font = bold
        ws.cell(row=summary_row, column=6, value=_fmt_hms(avg)).font = bold
    else:
        ws.cell(row=summary_row, column=1, value="(no sessions logged for this date)")

    for col, width in zip("ABCDEF", [6, 12, 14, 14, 16, 14]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_range_xlsx(
    camera: str,
    from_date: str,
    to_date: str,
    db_path: str = DEFAULT_DB_PATH,
) -> bytes:
    """One row per day in the range with that day's average wait + count."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference

    rows = daily_summary(camera, from_date, to_date, db_path=db_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "daily averages"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center")

    ws["A1"] = "Camera"
    ws["B1"] = camera
    ws["A2"] = "Range"
    ws["B2"] = f"{from_date} → {to_date}"
    ws["A1"].font = bold
    ws["A2"].font = bold

    headers = ["Date", "People", "Avg wait (seconds)", "Avg wait (mm:ss)", "Min wait", "Max wait"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = center

    for i, r in enumerate(rows, 1):
        ws.cell(row=4 + i, column=1, value=r["date"])
        ws.cell(row=4 + i, column=2, value=r["people"])
        ws.cell(row=4 + i, column=3, value=round(float(r["avg_wait"] or 0), 1))
        ws.cell(row=4 + i, column=4, value=_fmt_hms(float(r["avg_wait"] or 0)))
        ws.cell(row=4 + i, column=5, value=_fmt_hms(float(r["min_wait"] or 0)))
        ws.cell(row=4 + i, column=6, value=_fmt_hms(float(r["max_wait"] or 0)))

    if rows:
        chart = LineChart()
        chart.title = f"{camera}: daily average wait time (seconds)"
        chart.y_axis.title = "Average wait (s)"
        chart.x_axis.title = "Date"
        data_ref = Reference(ws, min_col=3, min_row=4, max_row=4 + len(rows))
        cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, f"H4")

    for col, width in zip("ABCDEF", [12, 10, 18, 16, 12, 12]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_monthly_xlsx(camera: str, year: int, db_path: str = DEFAULT_DB_PATH) -> bytes:
    """One row per month with that month's average wait, plus an embedded chart."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference

    rows = monthly_summary(camera, year, db_path=db_path)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    center = Alignment(horizontal="center")

    ws["A1"] = "Camera"
    ws["B1"] = camera
    ws["A2"] = "Year"
    ws["B2"] = year
    ws["A1"].font = bold
    ws["A2"].font = bold

    headers = ["Month", "People", "Avg wait (seconds)", "Avg wait (mm:ss)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = center

    for i, r in enumerate(rows, 1):
        ws.cell(row=4 + i, column=1, value=r["month"])
        ws.cell(row=4 + i, column=2, value=r["people"])
        ws.cell(row=4 + i, column=3, value=round(float(r["avg_wait"] or 0), 1))
        ws.cell(row=4 + i, column=4, value=_fmt_hms(float(r["avg_wait"] or 0)))

    if rows:
        chart = BarChart()
        chart.title = f"{camera}: monthly average wait time ({year})"
        chart.y_axis.title = "Average wait (s)"
        chart.x_axis.title = "Month"
        data_ref = Reference(ws, min_col=3, min_row=4, max_row=4 + len(rows))
        cats_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, "F4")

    for col, width in zip("ABCD", [10, 10, 18, 16]):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
